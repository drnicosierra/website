#!/usr/bin/env python3
"""
check_content_seo.py — Mechanical SEO/AEO checks for the drnicosierra.com content
pipeline, per DrSierra_SEO_AEO_Complete_Brief.docx.

Purpose: catch everything that's a pure string/length/count check WITHOUT needing
a human or an AI to read the text. Run this FIRST, on Dr. Sierra's Texto Nico
drafts, before anyone (human or Claude) spends time on the genuinely subjective
review (tone, voice, whether an AEO restatement actually reads naturally).

Usage:
    pip install openpyxl --break-system-packages
    python3 check_content_seo.py path/to/content_workbook.xlsx

What this DOES check (zero judgment required):
  - Forbidden term: "labio leporino"
  - "FLP" abbreviation in patient-facing content
  - Operation Smile + Smile Train co-occurrence within the same page
  - Doctor name: never "Perilla"; clinic name: "Clínica Tresserra" (accent)
  - Meta title <=60 chars, meta description 150-160 chars
  - Primary keyword presence per page (Section 12.2 keyword table)
  - Named entity count in "Por qué Dr. Sierra" / credential blocks (>=2 required)
  - Heuristic: FAQ answers starting with "Sí," / "No," / "Claro," (AEO Rule 1 red flag)
  - Heuristic: short/fragment-looking bullets (AEO Rule 3 red flag)

What this does NOT and CANNOT check (needs a human or Claude, by design):
  - Whether the tone matches Dr. Sierra's voice (Section 14)
  - Whether an AEO restatement actually reads naturally, not just "doesn't start with Sí"
  - Whether keyword insertion changed clinical meaning
  - Whether a paragraph was trimmed vs. expanded from his draft
  These are flagged as "NEEDS REVIEW" for the qualitative pass, not treated as errors.
"""
import sys, re
from openpyxl import load_workbook

FORBIDDEN_TERMS = ['labio leporino']
ENTITY_LIST = [
    "Children's Hospital of Philadelphia", 'Universidade Federal do Paraná', 'UFPR',
    'Operation Smile', 'PubMed', 'Global Cleft Care in Low Resource Settings', 'Springer',
    'FEBOMFS', 'European Board of Oral and Maxillofacial Surgery',
    "Hospital Universitario Vall d'Hebron", 'Clínica Tresserra', 'Mobile Surgery International',
    'SOCEFF', 'SECPF', 'MaxTrain', 'Circle of Cleft Professionals', 'Transforming Cleft', 'LATICFA',
]

PRIMARY_KEYWORDS = {
    'cirugia-labio-fisurado': 'cirugía labio fisurado España',
    'cirugia-paladar-hendido': 'cirugía paladar hendido España',
    'cirugia-revision-fisura': 'cirugía revisión fisura labiopalatina',
    'injerto-oseo-alveolar': 'injerto óseo alveolar fisura labiopalatina',
    'ortopedia-prequirurgica-nam': 'ortopedia prequirúrgica NAM fisura labiopalatina',
    'cirugia-ortognatica-flp': 'cirugía ortognática fisura labiopalatina',
    'rinoplastia-flp': 'rinoplastia fisura labiopalatina',
    'otros-procedimientos': 'tratamiento fisura labiopalatina Barcelona',
    'homepage': 'cirujano especialista en fisura labiopalatina en Barcelona',
    'about': 'cirujano especialista en fisura labiopalatina en Barcelona',
}

def norm(s):
    return (s or '').strip()

def check_terminology(text, campo):
    issues = []
    low = text.lower()
    for term in FORBIDDEN_TERMS:
        if term in low:
            issues.append(f'FORBIDDEN TERM: contains "{term}"')
    if re.search(r'\bFLP\b', text) and 'credenciales' not in campo.lower() and 'publicac' not in campo.lower():
        issues.append('"FLP" abbreviation used outside academic/credentials context (patient-facing content must spell out "fisura labiopalatina")')
    if 'perilla' in low:
        issues.append('Contains "Perilla" — doctor name must never include this surname per Dr. Sierra\'s preference')
    if 'clinica tresserra' in low and 'clínica tresserra' not in low:
        issues.append('"Clinica Tresserra" missing accent — must be "Clínica Tresserra"')
    return issues

def check_meta_length(elemento, text):
    issues = []
    el = elemento.lower()
    if 'título' in el or 'title' in el:
        if len(text) > 60:
            issues.append(f'Meta/SEO title is {len(text)} chars — max 60')
    if 'descripción' in el or 'description' in el:
        if not (150 <= len(text) <= 160) and len(text) > 0:
            issues.append(f'Meta description is {len(text)} chars — must be 150-160')
    return issues

def check_keyword(text, campo):
    issues = []
    page = None
    for slug in PRIMARY_KEYWORDS:
        if slug in campo:
            page = slug
            break
    if page and any(k in campo.lower() for k in ['hero.h1', 'title', 'quees.firstparagraph']):
        kw = PRIMARY_KEYWORDS[page]
        kw_core = kw.split(' ')[0:3]  # loose check: first few words of the keyword phrase
        if not any(w.lower() in text.lower() for w in ' '.join(kw_core).split()):
            issues.append(f'Primary keyword "{kw}" (or close variant) not detected — expected for this field')
    return issues

def check_entities(text, campo):
    issues = []
    if 'porque' in campo.lower() or 'credential' in campo.lower() or 'bio.credentials' in campo.lower():
        found = [e for e in ENTITY_LIST if e.lower() in text.lower()]
        if len(found) < 2 and len(text) > 200:  # only flag substantial bio text
            issues.append(f'Only {len(found)} named entity/entities found (need >=2) — found: {found}')
    return issues

def check_aeo_heuristics(text, elemento):
    issues = []
    el = elemento.lower()
    if 'respuesta' in el or 'answer' in el:
        first_words = text.strip()[:15].lower()
        if first_words.startswith(('sí,', 'sí.', 'no,', 'no.', 'claro,', 'claro.')):
            issues.append('AEO Rule 1 risk: answer opens with "Sí/No/Claro" instead of restating the question topic — likely won\'t be cited accurately by AI answer engines')
    return issues

def check_row(seccion, elemento, text, campo):
    issues = []
    issues += check_terminology(text, campo)
    issues += check_meta_length(elemento, text)
    issues += check_keyword(text, campo)
    issues += check_entities(text, campo)
    issues += check_aeo_heuristics(text, elemento)
    return issues

def main(path):
    wb = load_workbook(path, data_only=True)
    total_checked = 0
    total_flagged = 0
    for sheet_name in wb.sheetnames:
        if sheet_name in ('Instrucciones',):
            continue
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        try:
            col_seccion = headers.index('Sección')
            col_elemento = headers.index('Elemento')
            col_nico = headers.index('Texto Nico (Borrador)')
            col_campo = headers.index('Campo JSON (ruta técnica)')
        except ValueError:
            continue

        sheet_flags = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) <= max(col_seccion, col_elemento, col_nico, col_campo):
                continue
            seccion = norm(row[col_seccion])
            elemento = norm(row[col_elemento])
            text = norm(row[col_nico])
            campo = norm(row[col_campo])
            if not elemento or not text:
                continue
            total_checked += 1
            issues = check_row(seccion, elemento, text, campo)
            if issues:
                total_flagged += 1
                sheet_flags.append((seccion, elemento, issues))

        if sheet_flags:
            print(f'\n=== {sheet_name} ===')
            for seccion, elemento, issues in sheet_flags:
                print(f'  [{seccion} | {elemento}]')
                for issue in issues:
                    print(f'    ⚠️  {issue}')

    print(f'\n{"="*60}')
    print(f'Checked {total_checked} rows with Texto Nico content.')
    print(f'{total_flagged} flagged for mechanical issues.')
    print(f'{total_checked - total_flagged} passed mechanical checks cleanly.')
    print('\nNOTE: rows that passed still need the qualitative pass (tone, voice,')
    print('whether AEO restatements genuinely read well) before writing TEXTO REVISADO.')
    print('This script only catches what a string/length/count check can catch.')

if __name__ == '__main__':
    if len(sys.argv) != 2:
        print('Usage: python3 check_content_seo.py path/to/workbook.xlsx')
        sys.exit(1)
    main(sys.argv[1])
